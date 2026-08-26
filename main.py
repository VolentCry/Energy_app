"""Energy drink collection manager for Flet."""

from __future__ import annotations

import json
import colorsys
from io import BytesIO
import re
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path
from typing import Any
from uuid import uuid4

import flet as ft
from openpyxl import load_workbook


FIELDS = [
    "brand", "name", "flavor", "price", "discount_price", "manufacturer",
    "bottling_company", "caffeine_per_100ml", "taurine_per_100ml", "volume_l",
    "sugar_per_100ml", "purchase_place", "rating",
]
LABELS = {
    "brand": "Бренд", "name": "Название", "flavor": "Вкус", "price": "Стоимость",
    "discount_price": "Цена по скидке", "manufacturer": "Производитель",
    "bottling_company": "Компания розлива", "caffeine_per_100ml": "Кофеин, мг/100 мл",
    "taurine_per_100ml": "Таурин, мг/100 мл", "volume_l": "Объем, л",
    "sugar_per_100ml": "Сахар, г/100 мл", "purchase_place": "Место покупки",
    "rating": "Оценка",
}
DETAIL_FIELDS = {
    "flavor": "Вкус",
    "volume_l": "Объем",
    "price": "Стоимость",
    "discount_price": "Цена по скидке",
    "manufacturer": "Производитель",
    "bottling_company": "Компания розлива",
    "caffeine_per_100ml": "Кофеин",
    "taurine_per_100ml": "Таурин",
    "sugar_per_100ml": "Сахар",
    "purchase_place": "Место покупки",
}
DEFAULT_INFO_FIELDS = list(DETAIL_FIELDS)
COLLECTION_PAGE_SIZE = 50
class AppPreferences:
    """Хранит пользовательские настройки приложения в JSON-файле."""

    def __init__(self, settings_file: Path) -> None:
        self.settings_file = settings_file
        self.theme = "system"
        self.display = "list"
        self.info_fields = DEFAULT_INFO_FIELDS.copy()
        self.custom_info_fields: list[str] = []
        self.has_visited = False
        self.load()

    def load(self) -> None:
        """Загружает тему и признак предыдущего запуска из файла настроек."""
        if not self.settings_file.exists():
            return
        try:
            data = json.loads(self.settings_file.read_text(encoding="utf-8"))
            if data.get("theme") in {"light", "dark", "system"}:
                self.theme = data["theme"]
            if data.get("display") in {"list", "grid"}:
                self.display = data["display"]
            if isinstance(data.get("info_fields"), list):
                self.info_fields = [field for field in data["info_fields"] if field in DETAIL_FIELDS]
            if isinstance(data.get("custom_info_fields"), list):
                self.custom_info_fields = [clean(field) for field in data["custom_info_fields"] if clean(field)]
            self.has_visited = bool(data.get("has_visited", False))
        except (OSError, ValueError):
            pass

    def save(self) -> None:
        """Сохраняет пользовательские настройки на диске устройства."""
        self.settings_file.parent.mkdir(parents=True, exist_ok=True)
        self.settings_file.write_text(
            json.dumps({
                "theme": self.theme,
                "display": self.display,
                "info_fields": self.info_fields,
                "custom_info_fields": self.custom_info_fields,
                "has_visited": self.has_visited,
            }, indent=2),
            encoding="utf-8",
        )


def clean(value: Any) -> str:
    """Преобразует значение в строку и убирает пробелы по краям."""
    return "" if value is None else str(value).strip()


def number(value: Any, default: float | None = None) -> float | None:
    """Извлекает первое число из значения или возвращает значение по умолчанию."""
    text = clean(value).lower().replace(" ", "").replace(",", ".")
    if not text or text in {"-", "неуказан", "не указано", "безсахара", "отсутствует", "безтаурина"}:
        return default
    match = re.search(r"\d+(?:\.\d+)?", text)
    return float(match.group()) if match else default


def format_number(value: Any) -> str:
    """Преобразует число в компактную строку без лишних нулей."""
    parsed = number(value)
    return "" if parsed is None else f"{parsed:g}"


def empty_item() -> dict[str, Any]:
    """Создает пустую запись энергетика со значениями по умолчанию."""
    return {field: (5 if field == "rating" else "") for field in FIELDS} | {"card_color": "", "photo": "", "custom_fields": {}}


def theme_colors(workbook: Any) -> list[str]:
    """Извлекает цветовую схему темы Excel из XML-данных книги."""
    if not workbook.loaded_theme:
        return []
    root = ET.fromstring(workbook.loaded_theme)
    scheme = next((element for element in root.iter() if element.tag.endswith("clrScheme")), None)
    if scheme is None:
        return []
    colors = []
    for element in list(scheme):
        color = next(iter(element), None)
        value = color.attrib.get("lastClr") or color.attrib.get("val") if color is not None else None
        colors.append(f"#{value}" if value else "")
    return colors


def row_color(cell: Any, colors: list[str]) -> str:
    """Возвращает цвет заливки ячейки с учетом RGB, темы и оттенка."""
    fill = cell.fill
    if fill.fill_type != "solid":
        return ""
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        return f"#{color.rgb[-6:]}"
    if color.type == "theme" and color.theme is not None and color.theme < len(colors):
        base = colors[color.theme].lstrip("#")
        if len(base) != 6:
            return ""
        tint = color.tint or 0
        channels = [int(base[index:index + 2], 16) for index in range(0, 6, 2)]
        channels = [round(channel * (1 + tint) if tint < 0 else channel + (255 - channel) * tint) for channel in channels]
        return "#" + "".join(f"{max(0, min(255, channel)):02X}" for channel in channels)
    return ""


class CollectionStore:
    def __init__(self, data_file: Path) -> None:
        self.data_file = data_file
        self.photos_directory = data_file.parent / "photos"
        self.items: list[dict[str, Any]] = []
        self.load()

    def load(self) -> None:
        """Загружает сохраненную коллекцию из JSON-файла приложения."""
        source = self.data_file
        if source.exists():
            try:
                raw = json.loads(source.read_text(encoding="utf-8"))
                self.items = [self.normalize(item) for item in raw]
            except (OSError, ValueError):
                self.items = []

    def save(self) -> None:
        """Сохраняет текущие записи коллекции в JSON-файл."""
        self.data_file.parent.mkdir(parents=True, exist_ok=True)
        self.data_file.write_text(json.dumps(self.items, ensure_ascii=False, indent=2), encoding="utf-8")

    def normalize(self, item: dict[str, Any]) -> dict[str, Any]:
        """Заполняет отсутствующие поля записи и нормализует ее оценку."""
        result = empty_item()
        result.update({field: item.get(field, result[field]) for field in FIELDS})
        result["card_color"] = clean(item.get("card_color"))
        result["photo"] = clean(item.get("photo"))
        result["custom_fields"] = {
            clean(key): clean(value)
            for key, value in item.get("custom_fields", {}).items()
            if clean(key)
        } if isinstance(item.get("custom_fields", {}), dict) else {}
        result["rating"] = max(1, min(10, int(number(result["rating"], 5) or 5)))
        return result

    def save_photo(self, content: bytes, suffix: str) -> str:
        """Сохраняет изображение в каталоге фотографий и возвращает его имя."""
        self.photos_directory.mkdir(parents=True, exist_ok=True)
        filename = f"{uuid4().hex}{suffix.lower()}"
        (self.photos_directory / filename).write_bytes(content)
        return filename

    def photo_path(self, item: dict[str, Any]) -> Path | None:
        """Возвращает путь к фотографии записи, если файл существует."""
        filename = clean(item.get("photo"))
        if not filename:
            return None
        path = self.photos_directory / Path(filename).name
        return path if path.is_file() else None

    def excel_headers(self, source: Path | bytes) -> list[str]:
        """Возвращает непустые заголовки первого ряда Excel-файла."""
        workbook_source = BytesIO(source) if isinstance(source, bytes) else source
        workbook = load_workbook(workbook_source, read_only=True, data_only=True)
        sheet = workbook["Перечень"]
        return [clean(cell.value) for cell in next(sheet.iter_rows()) if clean(cell.value)]

    def import_excel(self, source: Path | bytes, selected_headers: set[str] | None = None) -> int:
        """Импортирует записи из листа «Перечень» Excel-файла и сохраняет их."""
        workbook_source = BytesIO(source) if isinstance(source, bytes) else source
        workbook = load_workbook(workbook_source, read_only=True, data_only=True)
        sheet = workbook["Перечень"]
        colors = theme_colors(workbook)
        headers = [clean(cell.value).lower() for cell in next(sheet.iter_rows())]
        aliases = {
            "бренд": "brand", "название": "name", "вкус": "flavor", "стоимость": "price",
            "стоимость по скидке": "discount_price", "производитель": "manufacturer",
            "компания розлива": "bottling_company", "кофеин(на 100 мл)": "caffeine_per_100ml",
            "таурин(на 100 мл)": "taurine_per_100ml", "объём (л)": "volume_l",
            "объем (л)": "volume_l", "сахар(на 100 мл)": "sugar_per_100ml",
            "место покупки": "purchase_place", "оценка": "rating",
        }
        indexes = {
            aliases[header]: index
            for index, header in enumerate(headers)
            if header in aliases and (selected_headers is None or header in selected_headers)
        }
        imported: list[dict[str, Any]] = []
        numeric_fields = {"price", "discount_price", "caffeine_per_100ml", "taurine_per_100ml", "volume_l", "sugar_per_100ml"}
        for cells in sheet.iter_rows(min_row=2):
            row = [cell.value for cell in cells]
            if not any(clean(value) for value in row):
                continue
            item = empty_item()
            item["card_color"] = row_color(cells[0], colors) if cells else ""
            for field, index in indexes.items():
                value = row[index] if index < len(row) else None
                item[field] = format_number(value) if field in numeric_fields else clean(value)
            item["rating"] = max(1, min(10, int(number(item["rating"], 5) or 5)))
            imported.append(item)
        self.items = imported
        self.save()
        return len(imported)


def metric_card(title: str, value: str, icon: ft.IconData, color: str) -> ft.Card:
    """Создает карточку статистики с иконкой, подписью и значением."""
    return ft.Card(content=ft.Container(padding=16, content=ft.Row([
        ft.Icon(icon, size=30, color=color),
        ft.Column([ft.Text(title, size=12, color=ft.Colors.ON_SURFACE_VARIANT), ft.Text(value, size=22, weight=ft.FontWeight.BOLD)], spacing=2, tight=True),
    ], spacing=12)))


class CollectionScreen(ft.Column):
    def __init__(self, app: "EnergyApp") -> None:
        self.app = app
        self._mounted = False
        self.search = ft.TextField(
            label="Поиск",
            prefix_icon=ft.Icons.SEARCH,
            suffix=ft.IconButton(ft.Icons.CLEAR, tooltip="Очистить поиск", on_click=self.clear_search),
            on_change=self.refresh,
            expand=True,
            height=72,
        )
        self.sort = ft.Dropdown(label="Сортировка", value="brand", options=[
            ft.DropdownOption("brand", "Бренд"), ft.DropdownOption("rating", "Оценка"), ft.DropdownOption("price", "Цена"),
        ], on_select=self.refresh, expand=True, height=72)
        self.list_view = ft.ListView(expand=True, spacing=6, padding=ft.Padding.symmetric(horizontal=12, vertical=8))
        self.grid_view = ft.GridView(
            expand=True,
            max_extent=240,
            spacing=6,
            run_spacing=2,
            child_aspect_ratio=1.55,
            padding=ft.Padding.symmetric(horizontal=12, vertical=8),
        )
        self.filtered_items: list[dict[str, Any]] = []
        self.visible_count = COLLECTION_PAGE_SIZE
        super().__init__(expand=True, spacing=8, controls=[
            ft.Container(
                height=96,
                padding=ft.Padding.only(left=12, right=12, top=12),
                content=ft.Row([
                    self.search,
                    self.sort,
                ], spacing=8),
            ),
            self.list_view if self.app.preferences.display == "list" else self.grid_view,
        ])
        self.refresh()

    def refresh(self, _e: Any = None) -> None:
        """Фильтрует, сортирует и отображает первую страницу коллекции."""
        query = (self.search.value or "").lower().strip()
        items = [item for item in self.app.store.items if query in " ".join(str(item.get(field, "")) for field in ["brand", "name", "flavor", "purchase_place"]).lower()]
        key = self.sort.value or "brand"
        if key == "rating":
            items.sort(key=lambda item: int(item.get(key) or 0), reverse=True)
        elif key == "price":
            items.sort(key=lambda item: number(item.get(key)) or 0)
        else:
            items.sort(key=lambda item: clean(item.get(key)).lower())
        self.filtered_items = items
        self.visible_count = COLLECTION_PAGE_SIZE
        self.update_list()

    def clear_search(self, _e: Any = None) -> None:
        """Очищает строку поиска и показывает всю коллекцию."""
        self.search.value = ""
        self.refresh()

    def update_list(self) -> None:
        """Обновляет видимые карточки и добавляет кнопку следующей страницы."""
        visible_items = self.filtered_items[:self.visible_count]
        controls: list[ft.Control] = [self.item_tile(item) for item in visible_items]
        if self.visible_count < len(self.filtered_items):
            controls.append(ft.Container(
                alignment=ft.Alignment.CENTER,
                padding=ft.Padding.only(top=4, bottom=12),
                content=ft.OutlinedButton(
                    f"Показать еще ({len(self.filtered_items) - self.visible_count})",
                    icon=ft.Icons.EXPAND_MORE,
                    on_click=self.show_more,
                ),
            ))
        view = self.list_view if self.app.preferences.display == "list" else self.grid_view
        view.controls = controls
        if self._mounted:
            self.update()

    def show_more(self, _e: Any = None) -> None:
        """Показывает следующую страницу найденных записей."""
        self.visible_count += COLLECTION_PAGE_SIZE
        self.update_list()

    def set_display(self, display: str) -> None:
        """Переключает список или сетку и обновляет отображение коллекции."""
        self.app.preferences.display = display if display in {"list", "grid"} else "list"
        self.app.preferences.save()
        self.controls[1] = self.list_view if self.app.preferences.display == "list" else self.grid_view
        self.update_list()

    def did_mount(self) -> None:
        """Отмечает экран как подключенный к дереву элементов Flet."""
        self._mounted = True

    def item_tile(self, item: dict[str, Any]) -> ft.Control:
        """Создает элемент списка с данными энергетика и действиями пользователя."""
        if self.app.preferences.display == "grid":
            return self.grid_item_tile(item)
        name = clean(item["name"]) or "Без названия"
        details = " · ".join(value for value in [item["flavor"], f"{item['volume_l']} л" if item["volume_l"] else "", item["purchase_place"]] if value)
        rating = int(item["rating"] or 0)
        stars = "★" * rating + "☆" * (10 - rating)
        rating_color = self.rating_color()
        accent_color = item.get("card_color") or ft.Colors.AMBER_700
        photo = self.app.store.photo_path(item)
        tile = ft.ListTile(
            leading=self.photo_or_avatar(item, photo, accent_color),
            title=ft.Text(name, size=16, weight=ft.FontWeight.BOLD, max_lines=2),
            subtitle=ft.Column([
                ft.Text(details or "Без дополнительных данных", size=13, max_lines=1, overflow=ft.TextOverflow.ELLIPSIS),
                ft.Row([
                    ft.Text(stars, size=12, color=rating_color),
                    ft.Text(f"{rating}/10", size=12, color=ft.Colors.ON_SURFACE_VARIANT),
                ], spacing=6),
            ], spacing=3, tight=True),
            trailing=ft.IconButton(ft.Icons.EDIT, tooltip="Изменить", on_click=lambda _e, value=item: self.app.open_form(value)),
            on_click=lambda _e, value=item: self.app.show_details(value),
            content_padding=ft.Padding.symmetric(horizontal=12, vertical=10),
        )
        card = ft.Card(
            elevation=3,
            content=ft.Row([
                ft.Container(width=5, height=102, bgcolor=accent_color, border_radius=ft.BorderRadius.only(top_left=8, bottom_left=8)),
                ft.Container(expand=True, content=tile),
            ], spacing=0),
        )
        return ft.Dismissible(
            content=card,
            background=ft.Container(
                alignment=ft.Alignment.CENTER_LEFT,
                padding=ft.Padding.only(left=24),
                bgcolor=ft.Colors.ERROR,
                content=ft.Icon(ft.Icons.DELETE, color=ft.Colors.WHITE, size=28),
            ),
            secondary_background=ft.Container(
                alignment=ft.Alignment.CENTER_RIGHT,
                padding=ft.Padding.only(right=24),
                bgcolor=ft.Colors.ERROR,
                content=ft.Icon(ft.Icons.DELETE, color=ft.Colors.WHITE, size=28),
            ),
            on_confirm_dismiss=lambda _e, value=item: self.confirm_swipe(value),
        )

    def photo_or_avatar(self, item: dict[str, Any], photo: Path | None, color: str) -> ft.Control:
        """Создает миниатюру фотографии или аватар с первой буквой бренда."""
        if photo:
            return ft.Image(src=str(photo), width=52, height=52, fit=ft.BoxFit.COVER, border_radius=26)
        return ft.CircleAvatar(
            content=ft.Text((item["brand"] or "?")[:1].upper(), weight=ft.FontWeight.BOLD),
            bgcolor=color,
            color=ft.Colors.WHITE,
        )

    def grid_item_tile(self, item: dict[str, Any]) -> ft.Control:
        """Создает компактную карточку записи для режима сетки."""
        name = clean(item["name"]) or "Без названия"
        details = " · ".join(value for value in [
            item["flavor"],
            f"{item['volume_l']} л" if item["volume_l"] else "",
            item["purchase_place"],
        ] if value)
        rating = int(item["rating"] or 0)
        stars = "★" * rating + "☆" * (10 - rating)
        rating_color = self.rating_color()
        accent_color = item.get("card_color") or ft.Colors.AMBER_700
        photo = self.app.store.photo_path(item)
        card = ft.Card(
            elevation=3,
            content=ft.Container(
                height=152,
                clip_behavior=ft.ClipBehavior.HARD_EDGE,
                padding=ft.Padding.symmetric(horizontal=10, vertical=8),
                content=ft.Column([
                    ft.Row([
                        self.photo_or_avatar(item, photo, accent_color),
                        ft.Container(expand=True),
                        ft.IconButton(
                            ft.Icons.EDIT,
                            icon_size=20,
                            tooltip="Изменить",
                            on_click=lambda _e, value=item: self.app.open_form(value),
                        ),
                    ], spacing=4),
                    ft.Text(
                        name,
                        size=15,
                        weight=ft.FontWeight.BOLD,
                        max_lines=2,
                        overflow=ft.TextOverflow.ELLIPSIS,
                    ),
                    ft.Text(
                        details or "Без дополнительных данных",
                        size=12,
                        max_lines=1,
                        overflow=ft.TextOverflow.ELLIPSIS,
                    ),
                    ft.Row([
                        ft.Text(stars, size=10, color=rating_color),
                        ft.Text(f"{rating}/10", size=11, color=ft.Colors.ON_SURFACE_VARIANT),
                    ], spacing=4),
                ], spacing=3, tight=True),
            ),
        )
        return ft.Dismissible(
            content=card,
            background=ft.Container(
                alignment=ft.Alignment.CENTER_LEFT,
                padding=ft.Padding.only(left=16),
                bgcolor=ft.Colors.ERROR,
                content=ft.Icon(ft.Icons.DELETE, color=ft.Colors.WHITE, size=24),
            ),
            secondary_background=ft.Container(
                alignment=ft.Alignment.CENTER_RIGHT,
                padding=ft.Padding.only(right=16),
                bgcolor=ft.Colors.ERROR,
                content=ft.Icon(ft.Icons.DELETE, color=ft.Colors.WHITE, size=24),
            ),
            on_confirm_dismiss=lambda _e, value=item: self.confirm_swipe(value),
        )

    def rating_color(self) -> str:
        """Возвращает единый цвет звезд с учетом текущей темы приложения."""
        return ft.Colors.AMBER_300 if self.app.page.theme_mode == ft.ThemeMode.DARK else ft.Colors.AMBER_800

    def confirm_swipe(self, item: dict[str, Any]) -> bool:
        """Открывает подтверждение удаления после свайпа и отменяет автосмахивание."""
        self.confirm_delete(item)
        return False

    def confirm_delete(self, item: dict[str, Any]) -> None:
        """Показывает диалог подтверждения удаления записи."""
        name = f"{item['brand']} {item['name']}".strip() or "Без названия"
        dialog = ft.AlertDialog(
            title=ft.Text("Удалить энергетик?"),
            content=ft.Text(f"Запись «{name}» будет удалена из коллекции."),
            actions=[
                ft.TextButton("Отмена", on_click=lambda _e, value=item: self.cancel_delete(value)),
                ft.FilledButton("Удалить", icon=ft.Icons.DELETE, on_click=lambda _e: self.delete(item)),
            ],
        )
        self.app.page.show_dialog(dialog)

    def cancel_delete(self, item: dict[str, Any]) -> None:
        """Отменяет удаление и пересоздает карточку в исходном положении."""
        self.app.page.pop_dialog()
        self.refresh()

    def delete(self, item: dict[str, Any]) -> None:
        """Удаляет запись, сохраняет коллекцию и обновляет список."""
        self.app.page.pop_dialog()
        self.app.store.items.remove(item)
        self.app.store.save()
        self.refresh()


class StatsScreen(ft.Column):
    def __init__(self, app: "EnergyApp") -> None:
        self.app = app
        self._mounted = False
        super().__init__(expand=True, scroll=ft.ScrollMode.AUTO, spacing=12, controls=[])
        self.refresh()

    def refresh(self, _e: Any = None) -> None:
        """Пересчитывает показатели и обновляет экран статистики."""
        items = self.app.store.items
        total_price = sum(number(item["discount_price"]) or number(item["price"]) or 0 for item in items)
        total_liters = sum(number(item["volume_l"]) or 0 for item in items)
        brands = Counter(item["brand"] for item in items if item["brand"])
        leaderboard = []
        for place, (brand, count) in enumerate(brands.most_common(5), start=1):
            leaderboard.append(ft.Row([
                ft.Container(width=32, content=ft.Text(f"{place}.", color=ft.Colors.ON_SURFACE_VARIANT)),
                ft.Text(brand, expand=True, max_lines=1, overflow=ft.TextOverflow.ELLIPSIS),
                ft.Text(str(count), width=48, text_align=ft.TextAlign.RIGHT, weight=ft.FontWeight.BOLD),
            ], spacing=8))
        if not leaderboard:
            leaderboard.append(ft.Text("Пока нет данных", color=ft.Colors.ON_SURFACE_VARIANT))
        average = sum(int(item["rating"] or 0) for item in items) / len(items) if items else 0
        metrics = [
            metric_card("Всего банок", str(len(items)), ft.Icons.INVENTORY_2, ft.Colors.AMBER_700),
            metric_card("Объем", f"{total_liters:.3g} л", ft.Icons.LOCAL_DRINK, ft.Colors.TEAL_600),
            metric_card("Потрачено", f"{total_price:.0f} ₽", ft.Icons.PAYMENTS_OUTLINED, ft.Colors.GREEN_600),
            metric_card("Средняя оценка", f"{average:.1f}/10", ft.Icons.STAR, ft.Colors.ORANGE_600),
        ]
        metric_cells = [ft.Container(expand=1, height=100, content=card) for card in metrics]
        self.controls = [
            ft.Container(padding=ft.Padding.only(left=12, top=14), content=ft.Text("Обзор коллекции", size=22, weight=ft.FontWeight.BOLD)),
            ft.Row(metric_cells[:2], spacing=8),
            ft.Row(metric_cells[2:], spacing=8),
            ft.Card(content=ft.Container(padding=16, content=ft.Column([
                ft.Container(
                    alignment=ft.Alignment.CENTER,
                    content=ft.Text("Популярные бренды", weight=ft.FontWeight.BOLD),
                ),
                ft.Row([
                    ft.Container(width=32),
                    ft.Text("Бренд", expand=True, size=12, color=ft.Colors.ON_SURFACE_VARIANT),
                    ft.Text("Кол-во", width=48, size=12, color=ft.Colors.ON_SURFACE_VARIANT, text_align=ft.TextAlign.RIGHT),
                ], spacing=8),
                ft.Divider(height=1),
                *leaderboard,
            ], spacing=8, tight=True))),
        ]
        if self._mounted:
            self.update()

    def did_mount(self) -> None:
        """Отмечает экран статистики как подключенный к дереву элементов Flet."""
        self._mounted = True


class FormScreen(ft.Column):
    def __init__(self, app: "EnergyApp", item: dict[str, Any] | None) -> None:
        self.app = app
        self.item = item
        self.photo_name = clean(item.get("photo")) if item else ""
        self.card_color = clean(item.get("card_color")) if item else ft.Colors.ORANGE_700
        self.file_picker = ft.FilePicker()
        app.page.services.append(self.file_picker)
        self.fields: dict[str, ft.TextField] = {}
        self.custom_fields: dict[str, ft.TextField] = {}
        visible_text_fields = [
            "brand",
            "name",
            *[field for field in DETAIL_FIELDS if field in self.app.preferences.info_fields],
        ]
        controls: list[ft.Control] = []
        numeric_fields = {"price", "discount_price", "caffeine_per_100ml", "taurine_per_100ml", "volume_l", "sugar_per_100ml"}
        for field in visible_text_fields:
            self.fields[field] = ft.TextField(label=LABELS[field], value=clean(item[field]) if item else "", keyboard_type=ft.KeyboardType.NUMBER if field in numeric_fields else ft.KeyboardType.TEXT, expand=True)
            controls.append(self.fields[field])
        self.rating = ft.Slider(min=1, max=10, divisions=9, value=float(item["rating"]) if item else 5, label="{value}", on_change=self.rating_changed)
        self.rating_text = ft.Text(f"Оценка: {int(self.rating.value)}/10", weight=ft.FontWeight.BOLD)
        self.photo_status = ft.Text(
            "Фото прикреплено" if self.photo_name else "Фото не выбрано",
            color=ft.Colors.ON_SURFACE_VARIANT,
        )
        self.remove_photo_button = ft.OutlinedButton(
            "Удалить фото",
            icon=ft.Icons.DELETE_OUTLINE,
            disabled=not bool(self.photo_name),
            on_click=self.remove_photo,
        )
        self.color_button = ft.FilledButton(
            "Выбрать цвет карточки",
            icon=ft.Icons.PALETTE,
            on_click=self.open_color_picker,
        )
        self.color_preview = ft.Container(width=42, height=42, bgcolor=self.card_color, border_radius=21)
        custom_values = item.get("custom_fields", {}) if item else {}
        custom_controls = []
        for label in self.app.preferences.custom_info_fields:
            field = ft.TextField(label=label, value=clean(custom_values.get(label)), expand=True)
            self.custom_fields[label] = field
            custom_controls.append(field)
        super().__init__(expand=True, scroll=ft.ScrollMode.AUTO, spacing=10, controls=[
            ft.Container(padding=ft.Padding.only(left=12, top=12), content=ft.Text("Редактирование" if item else "Новый энергетик", size=22, weight=ft.FontWeight.BOLD)),
            ft.Container(
                padding=ft.Padding.symmetric(horizontal=16),
                content=ft.ResponsiveRow(controls, spacing=12, run_spacing=12),
            ),
            ft.Container(
                padding=ft.Padding.only(left=16, right=16),
                content=ft.Column([
                    ft.Row([
                        ft.FilledButton("Выбрать фото", icon=ft.Icons.ADD_A_PHOTO, on_click=self.pick_photo),
                        self.remove_photo_button,
                    ], spacing=8),
                    self.photo_status,
                ], spacing=4),
            ),
            ft.Container(
                padding=ft.Padding.symmetric(horizontal=16),
                content=ft.Row([self.color_button, self.color_preview], spacing=10),
            ),
            ft.Container(
                padding=ft.Padding.symmetric(horizontal=16),
                content=ft.Column(custom_controls, spacing=8),
            ) if custom_controls else ft.Container(),
            ft.Container(padding=ft.Padding.only(left=16), content=self.rating_text),
            self.rating,
            ft.Container(
                padding=ft.Padding.only(right=16, bottom=16),
                content=ft.Row([
                    ft.OutlinedButton("Отмена", icon=ft.Icons.CLOSE, on_click=lambda _e: self.app.show_collection()),
                    ft.FilledButton("Сохранить", icon=ft.Icons.SAVE_OUTLINED, on_click=self.save),
                ], alignment=ft.MainAxisAlignment.END),
            ),
        ])

    def rating_changed(self, _e: Any) -> None:
        """Обновляет текстовое отображение выбранной оценки."""
        self.rating_text.value = f"Оценка: {int(self.rating.value)}/10"
        self.rating_text.update()

    async def pick_photo(self, _e: Any) -> None:
        """Открывает выбор изображения и сохраняет выбранное фото после выбора."""
        files = await self.file_picker.pick_files(
            dialog_title="Выберите фотографию",
            allowed_extensions=["jpg", "jpeg", "png", "webp"],
            allow_multiple=False,
            with_data=True,
        )
        if not files:
            return
        selected = files[0]
        content = selected.bytes
        if not content and selected.path:
            content = Path(selected.path).read_bytes()
        if not content:
            return
        suffix = Path(selected.name).suffix or ".jpg"
        self.photo_name = self.app.store.save_photo(content, suffix)
        self.remove_photo_button.disabled = False
        self.photo_status.value = "Фото прикреплено"
        self.photo_status.update()
        self.remove_photo_button.update()

    def remove_photo(self, _e: Any) -> None:
        """Удаляет фотографию из текущей записи при сохранении формы."""
        self.photo_name = ""
        self.remove_photo_button.disabled = True
        self.photo_status.value = "Фото не выбрано"
        self.photo_status.update()
        self.remove_photo_button.update()

    def open_color_picker(self, _e: Any) -> None:
        """Открывает палитру для свободного выбора оттенка карточки."""
        color = self.card_color if self.card_color.startswith("#") else "#F57C00"
        red = int(color[1:3], 16) / 255
        green = int(color[3:5], 16) / 255
        blue = int(color[5:7], 16) / 255
        hue, saturation, value = colorsys.rgb_to_hsv(red, green, blue)
        hue_slider = ft.Slider(min=0, max=360, divisions=360, value=hue * 360, label="{value}°")
        saturation_slider = ft.Slider(min=0, max=100, divisions=100, value=saturation * 100, label="{value}%")
        value_slider = ft.Slider(min=0, max=100, divisions=100, value=value * 100, label="{value}%")
        preview = ft.Container(width=180, height=58, bgcolor=color, border_radius=8)

        def update_color(_event: Any = None) -> None:
            red, green, blue = colorsys.hsv_to_rgb(
                hue_slider.value / 360,
                saturation_slider.value / 100,
                value_slider.value / 100,
            )
            preview.bgcolor = "#" + "".join(f"{round(channel * 255):02X}" for channel in (red, green, blue))
            preview.update()

        hue_slider.on_change = update_color
        saturation_slider.on_change = update_color
        value_slider.on_change = update_color

        def save_color(_event: Any) -> None:
            self.card_color = preview.bgcolor
            self.color_preview.bgcolor = self.card_color
            self.color_preview.update()
            self.app.page.pop_dialog()

        dialog = ft.AlertDialog(
            title=ft.Text("Цвет карточки"),
            content=ft.Column([
                preview,
                ft.Text("Оттенок"), hue_slider,
                ft.Text("Насыщенность"), saturation_slider,
                ft.Text("Яркость"), value_slider,
            ], tight=True, spacing=4),
            actions=[
                ft.TextButton("Отмена", on_click=lambda _event: self.app.page.pop_dialog()),
                ft.FilledButton("Выбрать", on_click=save_color),
            ],
        )
        self.app.page.show_dialog(dialog)

    def save(self, _e: Any) -> None:
        """Собирает данные формы, создает или обновляет запись и сохраняет ее."""
        result = self.item.copy() if self.item else empty_item()
        result.update({field: control.value.strip() for field, control in self.fields.items()})
        result["rating"] = int(self.rating.value)
        result["card_color"] = self.card_color
        result["photo"] = self.photo_name
        result["custom_fields"] = {label: field.value.strip() for label, field in self.custom_fields.items()}
        if self.item is None:
            self.app.store.items.append(result)
        else:
            self.item.update(result)
        self.app.store.save()
        self.app.show_collection()


class SettingsScreen(ft.Column):
    def __init__(self, app: "EnergyApp") -> None:
        self.app = app
        self.dark_switch = ft.Switch(
            value=app.page.theme_mode == ft.ThemeMode.DARK,
            on_change=self.toggle_theme,
        )
        self.display_dropdown = ft.Dropdown(
            label="Отображение",
            value=app.preferences.display,
            options=[
                ft.DropdownOption("list", "Список", leading_icon=ft.Icons.VIEW_LIST),
                ft.DropdownOption("grid", "Сетка", leading_icon=ft.Icons.GRID_VIEW),
            ],
            on_select=self.change_display,
            expand=True,
        )
        self.file_picker = ft.FilePicker()
        app.page.services.append(self.file_picker)
        top_content = ft.Column([
            ft.Container(padding=ft.Padding.only(left=24, right=24, top=14), content=ft.Text("Настройки", size=22, weight=ft.FontWeight.BOLD)),
            ft.Card(content=ft.Container(padding=ft.Padding.symmetric(horizontal=24, vertical=14), content=ft.Row([
                ft.Text("Светлая тема", expand=True),
                self.dark_switch,
                ft.Text("Тёмная тема", expand=True, text_align=ft.TextAlign.RIGHT),
            ], spacing=12))),
            ft.Card(content=ft.Container(padding=ft.Padding.symmetric(horizontal=24, vertical=14), content=self.display_dropdown)),
            ft.Card(content=ft.Container(padding=ft.Padding.symmetric(horizontal=24, vertical=14), content=ft.Row([
                ft.Text("Информация", expand=True),
                ft.OutlinedButton("Подробнее...", on_click=self.show_information_settings),
            ]))),
            ft.Card(content=ft.Container(padding=ft.Padding.symmetric(horizontal=24, vertical=16), content=ft.Column([
                ft.ListTile(leading=ft.Icon(ft.Icons.FILE_UPLOAD), title=ft.Text("Импорт из Excel"), subtitle=ft.Text("Заменить коллекцию данными из файла .xlsx"), on_click=self.import_excel),
                ft.ListTile(leading=ft.Icon(ft.Icons.FILE_UPLOAD), title=ft.Text("Импорт из JSON"), subtitle=ft.Text("Загрузить сохраненную коллекцию из файла .json"), on_click=self.import_json),
                ft.ListTile(leading=ft.Icon(ft.Icons.SAVE_OUTLINED), title=ft.Text("Экспорт в JSON"), subtitle=ft.Text("Сохранить текущую коллекцию в файл .json"), on_click=self.export_json),
            ], spacing=4, tight=True))),
        ], spacing=12, tight=True)
        super().__init__(
            expand=True,
            scroll=ft.ScrollMode.AUTO,
            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
            controls=[
                top_content,
                ft.Container(
                    alignment=ft.Alignment.CENTER,
                    padding=ft.Padding.only(left=24, right=24, bottom=24),
                    content=ft.Text(
                        "Данные хранятся в каталоге приложения.",
                        size=11,
                        text_align=ft.TextAlign.CENTER,
                        color=ft.Colors.with_opacity(0.55, ft.Colors.ON_SURFACE),
                    ),
                ),
            ],
        )

    def toggle_theme(self, e: ft.ControlEvent) -> None:
        """Переключает светлую и темную тему приложения."""
        is_dark = bool(e.control.value)
        self.app.page.theme_mode = ft.ThemeMode.DARK if is_dark else ft.ThemeMode.LIGHT
        self.app.preferences.theme = "dark" if is_dark else "light"
        self.app.preferences.save()
        self.app.page.update()

    def change_display(self, e: ft.ControlEvent) -> None:
        """Сохраняет выбранный режим отображения и обновляет экран коллекции."""
        self.app.collection.set_display(e.control.value)
        self.app.page.update()

    def show_information_settings(self, _e: Any) -> None:
        """Открывает настройки стандартных и пользовательских полей информации."""
        checkboxes = [ft.Checkbox(
            label=label,
            data=field,
            value=field in self.app.preferences.info_fields,
        ) for field, label in DETAIL_FIELDS.items()]
        custom_fields = [ft.TextField(label=label, value=label, expand=True) for label in self.app.preferences.custom_info_fields]
        custom_column = ft.Column(spacing=6)

        def redraw_custom_fields(update_page: bool = True) -> None:
            custom_column.controls = [ft.Row([
                field,
                ft.IconButton(ft.Icons.DELETE_OUTLINE, tooltip="Удалить поле", on_click=lambda _e, control=field: remove_custom_field(control)),
            ], spacing=4) for field in custom_fields]
            if update_page:
                custom_column.update()

        def remove_custom_field(field: ft.TextField) -> None:
            if field in custom_fields:
                custom_fields.remove(field)
                redraw_custom_fields()

        new_field = ft.TextField(label="Название своего поля", expand=True)

        def add_custom_field(_e: Any) -> None:
            label = clean(new_field.value)
            if label and label not in [field.value for field in custom_fields]:
                custom_fields.append(ft.TextField(label=label, value=label, expand=True))
                new_field.value = ""
                redraw_custom_fields()

        def save_information(_e: Any) -> None:
            self.app.preferences.info_fields = [checkbox.data for checkbox in checkboxes if checkbox.value]
            self.app.preferences.custom_info_fields = [clean(field.value) for field in custom_fields if clean(field.value)]
            self.app.preferences.save()
            self.app.page.pop_dialog()

        dialog = ft.AlertDialog(
            title=ft.Text("Информация"),
            content=ft.Container(
                width=440,
                height=460,
                content=ft.Column([
                    ft.Text("Стандартные поля", weight=ft.FontWeight.BOLD),
                    ft.Column(checkboxes, spacing=2),
                    ft.Divider(),
                    ft.Text("Свои поля", weight=ft.FontWeight.BOLD),
                    custom_column,
                    ft.Row([new_field, ft.IconButton(ft.Icons.ADD, tooltip="Добавить поле", on_click=add_custom_field)], spacing=4),
                ], scroll=ft.ScrollMode.AUTO, spacing=6),
            ),
            actions=[
                ft.TextButton("Отмена", on_click=lambda _e: self.app.page.pop_dialog()),
                ft.FilledButton("Сохранить", icon=ft.Icons.SAVE_OUTLINED, on_click=save_information),
            ],
        )
        redraw_custom_fields(update_page=False)
        self.app.page.show_dialog(dialog)

    async def import_excel(self, _e: Any) -> None:
        """Открывает Excel-файл и предлагает выбрать импортируемые столбцы."""
        files = await self.file_picker.pick_files(
            dialog_title="Выберите Excel-файл",
            allowed_extensions=["xlsx"],
            allow_multiple=False,
            with_data=True,
        )
        if not files:
            return
        selected = files[0]
        source = selected.bytes if selected.bytes else Path(selected.path)
        try:
            headers = self.app.store.excel_headers(source)
            self.show_column_dialog(source, headers)
        except Exception as error:
            self.app.notify(f"Ошибка чтения заголовков: {error}")

    def show_column_dialog(self, source: Path | bytes, headers: list[str]) -> None:
        """Показывает чекбоксы заголовков и запускает импорт выбранных столбцов."""
        if not headers:
            self.app.notify("В Excel-файле не найдены заголовки столбцов")
            return
        checkboxes = [ft.Checkbox(label=header, value=True) for header in headers]

        def import_selected(_e: Any) -> None:
            selected_headers = {
                clean(checkbox.label).lower()
                for checkbox in checkboxes
                if checkbox.value
            }
            if not selected_headers:
                self.app.notify("Выберите хотя бы один столбец")
                return
            self.app.page.pop_dialog()
            self.finish_import(source, selected_headers)

        dialog = ft.AlertDialog(
            title=ft.Text("Выберите столбцы для импорта"),
            content=ft.Container(
                width=420,
                height=420,
                content=ft.Column(checkboxes, scroll=ft.ScrollMode.AUTO, spacing=2),
            ),
            actions=[
                ft.TextButton("Отмена", on_click=lambda _e: self.app.page.pop_dialog()),
                ft.FilledButton("Импортировать", icon=ft.Icons.FILE_UPLOAD, on_click=import_selected),
            ],
        )
        self.app.page.show_dialog(dialog)

    async def import_json(self, _e: Any) -> None:
        """Открывает JSON-файл, проверяет его структуру и импортирует записи."""
        files = await self.file_picker.pick_files(dialog_title="Выберите JSON-файл", allowed_extensions=["json"], allow_multiple=False, with_data=True)
        if not files:
            return
        try:
            selected = files[0]
            raw = selected.bytes.decode("utf-8") if selected.bytes else Path(selected.path).read_text(encoding="utf-8")
            data = json.loads(raw)
            if not isinstance(data, list) or not all(isinstance(item, dict) for item in data):
                raise ValueError("JSON должен содержать список записей")
            self.app.store.items = [self.app.store.normalize(item) for item in data]
            self.app.store.save()
            self.app.notify(f"Импортировано записей: {len(data)}")
            self.app.show_collection()
        except (OSError, UnicodeError, ValueError) as error:
            self.app.notify(f"Ошибка импорта JSON: {error}")

    async def export_json(self, _e: Any) -> None:
        """Преобразует коллекцию в JSON и предлагает сохранить ее в файл."""
        payload = json.dumps(self.app.store.items, ensure_ascii=False, indent=2).encode("utf-8")
        path = await self.file_picker.save_file(
            dialog_title="Сохранить коллекцию",
            file_name="collection.json",
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["json"],
            src_bytes=payload,
        )
        if path:
            self.app.notify("Коллекция экспортирована в JSON")

    def finish_import(self, source: Path | bytes, selected_headers: set[str] | None = None) -> None:
        """Запускает импорт Excel и показывает пользователю результат операции."""
        try:
            count = self.app.store.import_excel(source, selected_headers)
            self.app.notify(f"Импортировано записей: {count}")
            self.app.show_collection()
        except Exception as error:
            self.app.notify(f"Ошибка импорта: {error}")


class EnergyApp:
    titles = ["Коллекция", "Статистика", "Настройки"]

    def __init__(self, page: ft.Page, data_file: Path, preferences: AppPreferences) -> None:
        self.page = page
        self.preferences = preferences
        self.store = CollectionStore(data_file)
        self.page.on_close = self.save_on_close
        self._setup_page()
        self.collection = CollectionScreen(self)
        self.stats = StatsScreen(self)
        self.settings = SettingsScreen(self)
        self.content = ft.Container(expand=True, content=self.collection)
        self.title = ft.Text(self.titles[0], size=20, weight=ft.FontWeight.BOLD)
        self.add_button = ft.IconButton(ft.Icons.ADD, tooltip="Добавить", on_click=lambda _e: self.open_form(None))
        self.nav = ft.NavigationBar(selected_index=0, on_change=self.navigate, destinations=[
            ft.NavigationBarDestination(ft.Icons.INVENTORY_2, "Коллекция", ft.Icons.INVENTORY_2),
            ft.NavigationBarDestination(ft.Icons.BAR_CHART, "Статистика"),
            ft.NavigationBarDestination(ft.Icons.SETTINGS_OUTLINED, "Настройки", ft.Icons.SETTINGS),
        ])
        page.appbar = ft.AppBar(title=self.title, center_title=True, actions=[self.add_button])
        page.navigation_bar = self.nav
        page.add(self.content)

    def _setup_page(self) -> None:
        """Настраивает название, тему, отступы и размеры окна приложения."""
        self.page.title = "Моя коллекция энергетиков"
        self.page.padding = 0
        self.page.theme_mode = {
            "light": ft.ThemeMode.LIGHT,
            "dark": ft.ThemeMode.DARK,
        }.get(self.preferences.theme, ft.ThemeMode.SYSTEM)
        self.page.theme = ft.Theme(color_scheme_seed=ft.Colors.AMBER)
        self.page.dark_theme = ft.Theme(color_scheme_seed=ft.Colors.AMBER)
        try:
            self.page.window.width = 430
            self.page.window.height = 850
        except Exception:
            pass

    def navigate(self, e: ft.ControlEvent) -> None:
        """Переключает активный экран по выбранному пункту навигации."""
        index = e.control.selected_index
        screens = [self.collection, self.stats, self.settings]
        self.title.value = self.titles[index]
        self.content.content = screens[index]
        self.page.appbar.actions = [self.add_button] if index == 0 else []
        if index == 0:
            self.collection.refresh()
        elif index == 1:
            self.stats.refresh()
        self.page.update()

    def show_collection(self) -> None:
        """Показывает экран коллекции и выбирает его в нижней навигации."""
        self.nav.selected_index = 0
        self.navigate(type("Event", (), {"control": self.nav})())

    def open_form(self, item: dict[str, Any] | None) -> None:
        """Открывает форму создания новой или редактирования существующей записи."""
        self.nav.selected_index = 0
        self.title.value = "Энергетик"
        self.content.content = FormScreen(self, item)
        self.page.update()

    def show_details(self, item: dict[str, Any]) -> None:
        """Показывает подробности выбранного энергетика в диалоговом окне."""
        name = clean(item["name"]) or "Без названия"
        rating = int(item["rating"] or 0)
        rating_color = ft.Colors.AMBER_300 if self.page.theme_mode == ft.ThemeMode.DARK else ft.Colors.AMBER_800
        values = {
            "flavor": item["flavor"],
            "volume_l": f"{item['volume_l']} л" if item["volume_l"] else "",
            "price": item["price"],
            "discount_price": item["discount_price"],
            "manufacturer": item["manufacturer"],
            "bottling_company": item["bottling_company"],
            "caffeine_per_100ml": f"{item['caffeine_per_100ml']} мг/100 мл" if item["caffeine_per_100ml"] else "",
            "taurine_per_100ml": f"{item['taurine_per_100ml']} мг/100 мл" if item["taurine_per_100ml"] else "",
            "sugar_per_100ml": f"{item['sugar_per_100ml']} г/100 мл" if item["sugar_per_100ml"] else "",
            "purchase_place": item["purchase_place"],
        }
        rows = [(DETAIL_FIELDS[field], values[field]) for field in self.preferences.info_fields if field in values]
        custom_values = item.get("custom_fields", {})
        rows.extend((label, custom_values.get(label, "")) for label in self.preferences.custom_info_fields)
        detail_rows = [
            ft.Row([
                ft.Text(label, width=145, color=ft.Colors.ON_SURFACE_VARIANT),
                ft.Text(value or "Не указано", expand=True, text_align=ft.TextAlign.RIGHT),
            ], spacing=8)
            for label, value in rows
        ]
        dialog = ft.AlertDialog(
            title=ft.Row([
                ft.CircleAvatar(content=ft.Text((item["brand"] or "?")[:1].upper()), bgcolor=ft.Colors.AMBER_700, color=ft.Colors.WHITE),
                ft.Column([
                    ft.Text(name, weight=ft.FontWeight.BOLD, max_lines=2),
                    ft.Text(f"{'★' * rating}{'☆' * (10 - rating)}  {rating}/10", size=13, color=rating_color),
                ], spacing=2, tight=True, expand=True),
            ], spacing=12),
            content=ft.Column(detail_rows, spacing=8, scroll=ft.ScrollMode.AUTO),
            actions=[
                ft.TextButton("Закрыть", on_click=lambda _e: self.page.pop_dialog()),
                ft.FilledButton("Изменить", icon=ft.Icons.EDIT, on_click=lambda _e: (self.page.pop_dialog(), self.open_form(item))),
            ],
        )
        self.page.show_dialog(dialog)

    def notify(self, message: str) -> None:
        """Показывает короткое уведомление пользователю."""
        self.page.show_dialog(ft.SnackBar(ft.Text(message)))

    def save_on_close(self, _e: Any = None) -> None:
        """Сохраняет коллекцию при закрытии приложения."""
        self.store.save()
        self.preferences.save()

    def show_first_visit(self) -> None:
        """Показывает приветствие при первом запуске и отмечает запуск завершенным."""
        if self.preferences.has_visited:
            return
        self.preferences.has_visited = True
        self.preferences.save()
        self.page.show_dialog(ft.AlertDialog(
            title=ft.Text("Добро пожаловать"),
            content=ft.Text("Загрузите Excel-файл в разделе «Настройки», чтобы начать работу."),
            actions=[ft.TextButton("Понятно", on_click=lambda _e: self.page.pop_dialog())],
        ))


async def main(page: ft.Page) -> None:
    """Инициализирует хранилище, интерфейс приложения и первоначальные данные."""
    storage_paths = ft.StoragePaths()
    page.services.append(storage_paths)
    try:
        data_directory = await storage_paths.get_application_documents_directory()
    except Exception:
        data_directory = str(Path(__file__).parent)
    preferences = AppPreferences(Path(data_directory) / "settings.json")
    app = EnergyApp(page, Path(data_directory) / "collection.json", preferences)
    app.show_first_visit()


if __name__ == "__main__":
    ft.run(main)
